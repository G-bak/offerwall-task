from db.mysql import get_db, USE_DATABASE
from fastapi import APIRouter, Request, HTTPException, Depends, Body
from fastapi.templating import Jinja2Templates
from fastapi.responses import JSONResponse, PlainTextResponse, HTMLResponse, StreamingResponse
from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.exc import SQLAlchemyError, ProgrammingError
from pydantic import BaseModel
from services.utils import GUIDE_DIR
import os, json, re
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from io import BytesIO

router = APIRouter()
templates = Jinja2Templates(directory="templates")

class QueryInput(BaseModel):
    sql: str

class SQLSaveRequest(BaseModel):
    filename: str
    content: str

@router.get("/sql-workbench")
async def read_spec(request: Request):
    try:
        return templates.TemplateResponse("sql-workbench.html", {"request": request})
    except Exception:
        raise HTTPException(status_code=500, detail="An unexpected error occurred")   
    
@router.post("/sql-workbench/execute-query")
async def execute_query(query: QueryInput, db: AsyncSession = Depends(get_db)):
    try:
        sql = query.sql.strip()
        print(f"[QUERY 입력] SQL: \n{sql}")

        result = await db.execute(text(sql))

        # fetch 가능한 쿼리인지 확인
        if result.returns_rows:
            columns = result.keys()
            rows = result.fetchall()
            dict_rows = [dict(zip(columns, row)) for row in rows]

            if not dict_rows:
                null_row = {col: None for col in columns}
                return {"columns": list(columns), "rows": [null_row]}
            
            return {"columns": list(columns), "rows": dict_rows}
        else:
            affected = result.rowcount
            # 예: USE, INSERT, UPDATE, DDL 등
            return {"message": f"{affected} row(s) affected"}

    except ProgrammingError as pe:
        return JSONResponse(status_code=400, content={"error": str(pe.orig)})
    except SQLAlchemyError as e:
        print("[SQLAlchemyError]", e)
        return JSONResponse(status_code=400, content={"error": str(e)})
    
@router.get("/sql-workbench/database-name")
async def get_database_name():
    return {"name": USE_DATABASE}
    
@router.get("/sql-workbench/table-list")
async def get_actual_tables(db: AsyncSession = Depends(get_db)):
    result = await db.execute(text("SHOW TABLES"))
    tables = [row[0] for row in result.fetchall()]
    return {"tables": tables}

@router.get("/sql-workbench/guide-list")
async def list_guide_files():
    try:
        files = [
            f for f in os.listdir(GUIDE_DIR)
            if os.path.isfile(os.path.join(GUIDE_DIR, f))
        ]
        return {"files": files}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    
@router.get("/sql-workbench/guide-files/{filename}")
async def read_guide_file(filename: str):
    file_path = os.path.join(GUIDE_DIR, filename)

    if not os.path.isfile(file_path):
        raise HTTPException(status_code=404, detail="File not found")

    try:
        ext = os.path.splitext(filename)[1].lower()
        with open(file_path, "r", encoding="utf-8") as f:
            if ext == ".json":
                data = json.load(f)
                return JSONResponse(content=data)
            else:
                content = f.read()
                return PlainTextResponse(content)
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    
@router.post("/sql-workbench/guide-submit")
async def handle_submit(data: dict = Body(...)):
    print("제출된 데이터:", data)
    return {"status": "ok"}    

@router.get("/sql-workbench/js-script")
async def get_tracking_script():
    file_path = GUIDE_DIR / "2_유입 트래킹 JS 스크립트.js"
    try:
        with open(file_path, "r", encoding="utf-8") as f:
            return PlainTextResponse(f.read())
    except FileNotFoundError:
        return PlainTextResponse("// 트래킹 스크립트 없음", status_code=404)
    
@router.post("/sql-workbench/export-excel")
async def export_excel(request: Request):
    data = await request.json()

    headers = data.get("headers")
    rows = data.get("rows")

    if not headers or not rows:
        return {"error": "엑셀로 내보낼 데이터가 없습니다."}

    wb = Workbook()
    ws = wb.active
    ws.title = "데이터 리포트"

    # ✅ 스타일 정의
    header_font = Font(bold=True, color="FFFFFF")  # 흰색 글씨
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # 파란 계열
    alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )

    # ✅ 헤더 작성
    for col_num, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.alignment = alignment
        cell.border = thin_border
        cell.fill = header_fill

    # ✅ 데이터 작성
    for row_num, row_data in enumerate(rows, start=2):
        for col_num, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.alignment = alignment
            cell.border = thin_border

    # ✅ 열 너비 자동 조정 (한글도 고려해 충분히 넓게)
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                length = len(str(cell.value))
                if any('\uac00' <= c <= '\ud7a3' for c in str(cell.value)):
                    length *= 2  # 한글 보정
                max_length = max(max_length, length)
        ws.column_dimensions[col_letter].width = max_length * 1.2

    # ✅ 엑셀 파일 메모리로 저장 후 반환
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=report.xlsx"}
    )

@router.post("/save-sql-json")
async def save_sql_file(data: SQLSaveRequest):
    os.makedirs(GUIDE_DIR, exist_ok=True)

    # 1. 현재 디렉토리의 기존 SQL/JSON/JS 파일 목록
    existing_files = [
        f for f in os.listdir(GUIDE_DIR)
        if re.match(r"^\d{1,3}_.+\.(sql|json|js)$", f)
    ]

    # 2. 최대 번호 추출
    max_index = 0
    for fname in existing_files:
        match = re.match(r"^(\d{1,3})_", fname)
        if match:
            num = int(match.group(1))
            max_index = max(max_index, num)

    # 3. 다음 순번
    next_index = max_index + 1
    index_prefix = f"{next_index:01d}" if next_index < 10 else f"{next_index:03d}"

    # 4. 안전한 파일 이름 변환 (한글 허용)
    safe_name = re.sub(r"[^\w가-힣\- ]+", "_", data.filename.strip())

    # 5. 파일명 구성 및 저장
    filename = f"{index_prefix}_{safe_name}.sql"
    filepath = os.path.join(GUIDE_DIR, filename)

    with open(filepath, "w", encoding="utf-8") as f:
        f.write(data.content.strip() + "\n")

    return {"status": "success", "filename": filename}